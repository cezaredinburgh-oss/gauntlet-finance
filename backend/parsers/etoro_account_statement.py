"""eToro official multi-sheet Account Statement (.xlsx) parser.

Primary sheet: **Account Activity**
Optional: Dividends (if rows present)

Holdings snapshots are intentionally ignored (would double-count vs Activity).
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any, Mapping
from uuid import UUID, uuid4

from dateutil import parser as date_parser
from openpyxl import load_workbook

from backend.common.money import parse_decimal_optional
from backend.common.timeutil import utc_now
from backend.parsers.base import (
    ParseResult,
    empty_to_none,
    open_lot_from_buy,
    placeholder_basis,
    resolve_account_id,
)
from backend.schema.models import (
    AssetClass,
    InvestmentEvent,
    InvestmentEventType,
    ParserKey,
    TradeSide,
)

# Details like "ARKF/USD", "SOL/USD", "BTC/USD"
_PAIR_RE = re.compile(r"^([A-Za-z0-9.\-]+)\s*/\s*[A-Za-z]{3,4}$")
# Closed Positions style "Coinbase Global Inc (COIN.RTH)"
_PAREN_TICKER_RE = re.compile(r"\(([A-Za-z0-9.\-]+)\)\s*$")


def is_etoro_account_statement_xlsx(file_bytes: bytes) -> bool:
    """True if bytes look like eToro multi-sheet account statement workbook."""
    if not file_bytes or file_bytes[:2] != b"PK":
        return False
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            names = {n.strip().lower() for n in wb.sheetnames}
        finally:
            wb.close()
    except Exception:  # noqa: BLE001
        return False
    return "account activity" in names


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def _is_blank_token(raw: Any) -> bool:
    s = _cell_str(raw)
    return s == "" or s == "-"


def _parse_etoro_datetime(raw: Any) -> datetime:
    """eToro uses day-first timestamps (e.g. 17/11/2021 14:30:03)."""
    if isinstance(raw, datetime):
        dt = raw
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    if isinstance(raw, date) and not isinstance(raw, datetime):
        return datetime(raw.year, raw.month, raw.day, tzinfo=timezone.utc)

    s = _cell_str(raw)
    if not s:
        raise ValueError("empty date")
    # Prefer day-first for slash dates
    if re.match(r"^\d{1,2}/\d{1,2}/\d{4}", s):
        dt = date_parser.parse(s, dayfirst=True)
    else:
        dt = date_parser.parse(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _parse_units(raw: Any) -> Decimal | None:
    if _is_blank_token(raw):
        return None
    return parse_decimal_optional(_cell_str(raw) if not isinstance(raw, (int, float, Decimal)) else raw)


def _parse_amount(raw: Any) -> Decimal | None:
    if _is_blank_token(raw):
        return None
    if isinstance(raw, (int, float, Decimal)):
        return Decimal(str(raw))
    return parse_decimal_optional(_cell_str(raw))


def _ticker_from_details(details: str | None) -> str | None:
    d = empty_to_none(details)
    if not d or d == "-":
        return None
    m = _PAIR_RE.match(d)
    if m:
        return _normalize_ticker(m.group(1))
    m2 = _PAREN_TICKER_RE.search(d)
    if m2:
        return _normalize_ticker(m2.group(1))
    # bare symbol
    if re.match(r"^[A-Za-z0-9.\-]{1,20}$", d):
        return _normalize_ticker(d)
    return None


def _normalize_ticker(raw: str) -> str:
    t = raw.strip().upper()
    # Strip exchange session suffixes e.g. COIN.RTH
    if "." in t:
        base, suffix = t.rsplit(".", 1)
        if suffix in {"RTH", "ETH", "USD", "US", "L", "O"} and len(base) >= 1:
            # Keep ETH the crypto ticker; only strip when base is longer equity ticker
            if suffix == "ETH" and base in {"", "W"}:
                return t
            if suffix in {"RTH", "US", "L", "O"}:
                return base
            if suffix == "USD" and base not in {"", "T"}:
                return base
    return t


def _map_asset_class(raw: Any) -> AssetClass | None:
    if _is_blank_token(raw):
        return None
    t = _cell_str(raw).lower()
    if "crypto" in t:
        return AssetClass.CRYPTO
    if "stock" in t:
        return AssetClass.STOCK
    if "etf" in t:
        return AssetClass.ETF
    if "cfd" in t:
        return AssetClass.OTHER
    if "cash" in t:
        return AssetClass.CASH
    return AssetClass.OTHER


def _map_activity_type(raw: str) -> InvestmentEventType | None:
    t = (raw or "").strip().lower()
    mapping = {
        "open position": InvestmentEventType.BUY,
        "position closed": InvestmentEventType.SELL,
        "close position": InvestmentEventType.SELL,
        "staking": InvestmentEventType.STAKING_REWARD,
        "staking reward": InvestmentEventType.STAKING_REWARD,
        "deposit": InvestmentEventType.DEPOSIT,
        "withdraw request": InvestmentEventType.WITHDRAWAL,
        "withdrawal": InvestmentEventType.WITHDRAWAL,
        "withdraw fee": InvestmentEventType.FEE,
        "deposit conversion fee": InvestmentEventType.FEE,
        "opening and closing spread": InvestmentEventType.FEE,
        "overnight fee": InvestmentEventType.FEE,
        "fee": InvestmentEventType.FEE,
        "commission": InvestmentEventType.FEE,
        "dividend": InvestmentEventType.DEPOSIT,  # no Dividend enum; cash credit
        "adjusted dividend": InvestmentEventType.DEPOSIT,
        "corporate action": InvestmentEventType.TRANSFER,
        "edit stop loss": None,  # ignore UI noise if present
        "edit take profit": None,
    }
    if t in mapping:
        return mapping[t]
    if "spread" in t or "fee" in t or "commission" in t:
        return InvestmentEventType.FEE
    if "deposit" in t:
        return InvestmentEventType.DEPOSIT
    if "withdraw" in t:
        return InvestmentEventType.WITHDRAWAL
    if "staking" in t:
        return InvestmentEventType.STAKING_REWARD
    if "open" in t and "position" in t:
        return InvestmentEventType.BUY
    if "closed" in t or ("close" in t and "position" in t):
        return InvestmentEventType.SELL
    return None


def _sheet_rows(wb, name: str) -> list[dict[str, Any]]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    keys = []
    for h in header:
        keys.append(_cell_str(h) if h is not None else "")
    out: list[dict[str, Any]] = []
    for row in rows_iter:
        if not row or not any(c is not None and _cell_str(c) not in {"", "-"} for c in row):
            continue
        d: dict[str, Any] = {}
        for i, key in enumerate(keys):
            if not key:
                continue
            d[key] = row[i] if i < len(row) else None
        out.append(d)
    return out


def _col(row: dict[str, Any], *names: str) -> Any:
    # Exact then case-insensitive
    for n in names:
        if n in row:
            return row[n]
    lower = {k.lower(): v for k, v in row.items()}
    for n in names:
        if n.lower() in lower:
            return lower[n.lower()]
    return None


def parse_etoro_account_statement_bytes(
    file_bytes: bytes,
    *,
    account_ids: Mapping[str, UUID],
    source_file_id: UUID | None = None,
    file_hash: str | None = None,
    now: datetime | None = None,
) -> ParseResult:
    """Parse official eToro Account Statement xlsx into investment events/lots."""
    if not is_etoro_account_statement_xlsx(file_bytes):
        raise ValueError(
            "not an eToro Account Statement xlsx (missing 'Account Activity' sheet)"
        )

    ts = now or utc_now()
    account_id = resolve_account_id(account_ids, currency="USD")
    currency = "USD"

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        # Currency from Account Summary if present
        for row in _sheet_rows(wb, "Account Summary"):
            # rows are Details/Name style; openpyxl dict uses first header only once
            pass
        # Re-read Account Summary more carefully for Currency
        if "Account Summary" in wb.sheetnames:
            for r in wb["Account Summary"].iter_rows(values_only=True):
                if r and _cell_str(r[0]).lower() == "currency" and len(r) > 1:
                    ccy = empty_to_none(_cell_str(r[1]))
                    if ccy:
                        currency = ccy.upper()
                    break

        activity = _sheet_rows(wb, "Account Activity")
        dividends = _sheet_rows(wb, "Dividends")
    finally:
        wb.close()

    events: list[InvestmentEvent] = []
    lots = []
    # Position ID -> (buy_event_id, lot_index) for fee capitalization
    by_position: dict[str, tuple[UUID, int]] = {}
    data_rows = 0
    skipped = 0

    for row in activity:
        data_rows += 1
        type_raw = _cell_str(_col(row, "Type"))
        event_type = _map_activity_type(type_raw)
        if event_type is None:
            skipped += 1
            continue

        try:
            event_dt = _parse_etoro_datetime(_col(row, "Date"))
        except ValueError:
            skipped += 1
            continue
        event_date = event_dt.date()

        details = empty_to_none(_cell_str(_col(row, "Details")))
        amount = _parse_amount(_col(row, "Amount"))
        units = _parse_units(_col(row, "Units / Contracts", "Units"))
        position_id = empty_to_none(_cell_str(_col(row, "Position ID")))
        if position_id == "-":
            position_id = None
        asset_class = _map_asset_class(_col(row, "Asset type", "Asset Type"))
        ticker = _ticker_from_details(details)

        event_id = uuid4()
        parent_event_id = None
        lot_id = None
        side = None
        fees = Decimal("0")
        price = None
        description = details or type_raw

        if event_type == InvestmentEventType.BUY:
            side = TradeSide.BUY
            if ticker and units is not None and units > 0:
                cost = abs(amount) if amount is not None else Decimal("0")
                if units > 0 and cost > 0:
                    price = (cost / units).quantize(Decimal("0.00000001"))
                lot = open_lot_from_buy(
                    account_id=account_id,
                    ticker=ticker,
                    asset_class=asset_class or AssetClass.OTHER,
                    source="eToro",
                    acquisition_date=event_date,
                    quantity=units,
                    cost_native=cost,
                    native_currency=currency,
                    open_event_id=event_id,
                    now=ts,
                    notes=f"position_id={position_id}" if position_id else details,
                )
                lots.append(lot)
                lot_id = lot.id
                if position_id:
                    by_position[position_id] = (event_id, len(lots) - 1)

        elif event_type == InvestmentEventType.STAKING_REWARD:
            side = TradeSide.BUY
            if ticker and units is not None and units > 0:
                cost = abs(amount) if amount is not None else Decimal("0")
                lot = open_lot_from_buy(
                    account_id=account_id,
                    ticker=ticker,
                    asset_class=asset_class or AssetClass.CRYPTO,
                    source="eToro",
                    acquisition_date=event_date,
                    quantity=units,
                    cost_native=cost,
                    native_currency=currency,
                    open_event_id=event_id,
                    now=ts,
                    notes="staking",
                )
                lots.append(lot)
                lot_id = lot.id

        elif event_type == InvestmentEventType.SELL:
            side = TradeSide.SELL

        elif event_type == InvestmentEventType.FEE:
            # Capitalize spread/fees into open lot when Position ID matches a buy in-file
            fee_abs = abs(amount) if amount is not None else Decimal("0")
            fees = fee_abs
            if position_id and position_id in by_position:
                parent_event_id, lot_idx = by_position[position_id]
                lot = lots[lot_idx]
                new_native = lot.cost_basis_native + fee_abs
                n, czk, usd = placeholder_basis(new_native, lot.native_currency)
                note = (lot.notes or "") + f"; +{type_raw} {fee_abs}"
                lots[lot_idx] = lot.model_copy(
                    update={
                        "cost_basis_native": n,
                        "cost_basis_czk": czk,
                        "cost_basis_usd": usd,
                        "updated_at": ts,
                        "notes": note.strip("; "),
                    }
                )
            # Prefer ticker from linked position details if fee details empty
            if not ticker and position_id and position_id in by_position:
                _, lot_idx = by_position[position_id]
                ticker = lots[lot_idx].ticker

        # external_id: stable-ish for dedupe (position + type + timestamp + amount)
        ext_parts = [
            "etoro",
            type_raw.replace(" ", "_"),
            position_id or "",
            event_dt.isoformat(),
            _cell_str(amount) if amount is not None else "",
            _cell_str(units) if units is not None else "",
            details or "",
        ]
        external_id = "|".join(ext_parts)

        value_usd = amount if currency == "USD" and amount is not None else None
        value_czk = amount if currency == "CZK" and amount is not None else None

        events.append(
            InvestmentEvent(
                id=event_id,
                account_id=account_id,
                event_type=event_type,
                event_date=event_date,
                event_datetime=event_dt,
                ticker=ticker,
                asset_class=asset_class,
                side=side,
                quantity=units,
                price_native=price,
                native_currency=currency,
                value_native=amount,
                fees_native=fees,
                value_usd=value_usd,
                value_czk=value_czk,
                lot_id=lot_id,
                parent_event_id=parent_event_id,
                source="eToro",
                description=description,
                original_description=details,
                external_id=external_id,
                source_file_id=source_file_id,
                original_file_hash=file_hash,
                created_at=ts,
                updated_at=ts,
            )
        )

    # Dividends sheet (cash credits)
    for row in dividends:
        data_rows += 1
        try:
            event_dt = _parse_etoro_datetime(
                _col(row, "Date of Payment", "Date")
            )
        except ValueError:
            skipped += 1
            continue
        amount = _parse_amount(
            _col(row, "Net Dividend Received (USD)", "Net dividends", "Amount")
        )
        instrument = empty_to_none(_cell_str(_col(row, "Instrument Name", "Details")))
        ticker = _ticker_from_details(instrument) if instrument else None
        position_id = empty_to_none(_cell_str(_col(row, "Position ID")))
        if position_id == "-":
            position_id = None
        event_id = uuid4()
        events.append(
            InvestmentEvent(
                id=event_id,
                account_id=account_id,
                event_type=InvestmentEventType.DEPOSIT,
                event_date=event_dt.date(),
                event_datetime=event_dt,
                ticker=ticker,
                asset_class=_map_asset_class(_col(row, "Type")),
                side=None,
                quantity=None,
                price_native=None,
                native_currency=currency,
                value_native=amount,
                fees_native=Decimal("0"),
                value_usd=amount if currency == "USD" else None,
                value_czk=None,
                lot_id=None,
                parent_event_id=None,
                source="eToro",
                description=instrument or "Dividend",
                original_description=instrument,
                external_id=f"etoro|dividend|{position_id or ''}|{event_dt.isoformat()}|{amount}",
                source_file_id=source_file_id,
                original_file_hash=file_hash,
                created_at=ts,
                updated_at=ts,
            )
        )

    return ParseResult(
        parser_key=ParserKey.ETORO_ACCOUNT_STATEMENT.value,
        institution="eToro",
        row_count=data_rows,
        investment_events=events,
        investment_lots=lots,
    )
